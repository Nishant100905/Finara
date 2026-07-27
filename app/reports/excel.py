"""
Excel report generator using openpyxl.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from .models import FinancialReport


class ExcelReportGenerator:
    """
    Generates Excel reports.
    """

    def generate(
        self,
        report: FinancialReport,
        output_path: str,
    ) -> str:
        """
        Generate an Excel workbook.

        Returns:
            Path to generated workbook.
        """

        output = Path(output_path)
        output.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        self._build_summary_sheet(
            summary_sheet,
            report,
        )

        self._build_profile_sheet(
            workbook,
            report,
        )

        self._build_sections_sheet(
            workbook,
            report,
        )

        self._build_raw_data_sheet(
            workbook,
            report,
        )

        workbook.save(output)

        return str(output)

    def _heading(
        self,
        sheet,
        row: int,
        text: str,
    ) -> None:

        cell = sheet.cell(
            row=row,
            column=1,
        )

        cell.value = text
        cell.font = Font(
            bold=True,
            size=14,
        )

    def _build_summary_sheet(
        self,
        sheet,
        report: FinancialReport,
    ) -> None:

        self._heading(
            sheet,
            1,
            report.title,
        )

        sheet["A3"] = "Generated For"
        sheet["B3"] = report.generated_for

        sheet["A4"] = "Report Date"
        sheet["B4"] = str(report.report_date)

        sheet["A5"] = "Version"
        sheet["B5"] = report.metadata.version

        row = 7

        sheet.cell(
            row=row,
            column=1,
        ).value = "Executive Summary"

        sheet.cell(
            row=row,
            column=1,
        ).font = Font(
            bold=True,
        )

        row += 1

        sheet.cell(
            row=row,
            column=1,
        ).value = report.executive_summary.overview

        row += 2

        self._write_list(
            sheet,
            row,
            "Strengths",
            report.executive_summary.strengths,
        )

        row += len(
            report.executive_summary.strengths
        ) + 2

        self._write_list(
            sheet,
            row,
            "Risks",
            report.executive_summary.risks,
        )

        row += len(
            report.executive_summary.risks
        ) + 2

        self._write_list(
            sheet,
            row,
            "Recommendations",
            report.executive_summary.recommendations,
        )

    def _build_profile_sheet(
        self,
        workbook: Workbook,
        report: FinancialReport,
    ) -> None:

        sheet = workbook.create_sheet(
            "Profile"
        )

        self._heading(
            sheet,
            1,
            "User Profile",
        )

        profile = report.raw_data.get(
            "profile",
            {},
        )

        row = 3

        for key, value in profile.items():

            sheet.cell(
                row=row,
                column=1,
            ).value = key

            sheet.cell(
                row=row,
                column=2,
            ).value = str(value)

            row += 1

    def _build_sections_sheet(
        self,
        workbook: Workbook,
        report: FinancialReport,
    ) -> None:

        sheet = workbook.create_sheet(
            "Report Sections"
        )

        row = 1

        for section in sorted(
            report.sections,
            key=lambda s: s.order,
        ):

            sheet.cell(
                row=row,
                column=1,
            ).value = section.title

            sheet.cell(
                row=row,
                column=1,
            ).font = Font(
                bold=True,
            )

            row += 1

            sheet.cell(
                row=row,
                column=1,
            ).value = section.content

            row += 2

    def _build_raw_data_sheet(
        self,
        workbook: Workbook,
        report: FinancialReport,
    ) -> None:

        sheet = workbook.create_sheet(
            "Raw Data"
        )

        self._heading(
            sheet,
            1,
            "Raw Financial Data",
        )

        row = 3

        for key, value in report.raw_data.items():

            sheet.cell(
                row=row,
                column=1,
            ).value = key

            sheet.cell(
                row=row,
                column=2,
            ).value = str(value)

            row += 1

    def _write_list(
        self,
        sheet,
        start_row: int,
        title: str,
        items: list[str],
    ) -> None:

        sheet.cell(
            row=start_row,
            column=1,
        ).value = title

        sheet.cell(
            row=start_row,
            column=1,
        ).font = Font(
            bold=True,
        )

        row = start_row + 1

        for item in items:

            sheet.cell(
                row=row,
                column=1,
            ).value = item

            row += 1


excel_generator = ExcelReportGenerator()